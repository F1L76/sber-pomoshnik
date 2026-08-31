#!/usr/bin/env python3
"""Пишет широту/долготу из coords.jsonl в тот же xlsx со списком КН."""
import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

HEADERS = ("кадастровый", "широта", "долгота", "тип")


def load_coords(jsonl_path):
    coords = {}
    path = Path(jsonl_path)
    if not path.exists():
        return coords
    with path.open(encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            try:
                row = json.loads(line)
            except json.JSONDecodeError:
                continue
            kn = str(row.get("kn") or row.get("cadastralNumber") or "").strip()
            if not kn or not row.get("ok"):
                continue
            lat = row.get("lat")
            lon = row.get("lon")
            if lat is None or lon is None:
                continue
            coords[kn] = (round(float(lat), 6), round(float(lon), 6), str(row.get("t") or ""))
    return coords


def write_coords(xlsx_path, jsonl_path):
    coords = load_coords(jsonl_path)
    wb = load_workbook(xlsx_path)
    ws = wb.active
    for col, title in enumerate(HEADERS, 1):
        ws.cell(1, col, title)
    filled = 0
    for row_i in range(2, ws.max_row + 1):
        kn = str(ws.cell(row_i, 1).value or "").strip()
        hit = coords.get(kn)
        if not hit:
            continue
        lat, lon, kind = hit
        ws.cell(row_i, 2, lat)
        ws.cell(row_i, 3, lon)
        ws.cell(row_i, 4, kind)
        filled += 1
    wb.save(xlsx_path)
    return filled, len(coords)


def _selfcheck():
    import tempfile

    tmp = Path(tempfile.mkdtemp()) / "list.xlsx"
    jl = tmp.with_suffix(".jsonl")
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "кадастровый"
    ws["A2"] = "77:01:0001008:3459"
    ws["A3"] = "00:00:0000000:1"
    wb.save(tmp)
    jl.write_text(
        json.dumps({"kn": "77:01:0001008:3459", "ok": True, "lat": 55.757323261, "lon": 37.617424424, "t": "Помещения"})
        + "\n",
        encoding="utf-8",
    )
    filled, _ = write_coords(tmp, jl)
    wb2 = load_workbook(tmp, data_only=True)
    ws2 = wb2.active
    assert filled == 1, filled
    assert ws2["B2"].value == 55.757323
    assert ws2["C2"].value == 37.617424
    assert ws2["D2"].value == "Помещения"
    assert ws2["B3"].value is None
    tmp.unlink()
    jl.unlink()
    print("ok: write-coords-to-xlsx")


if __name__ == "__main__":
    if len(sys.argv) == 2 and sys.argv[1] == "--check":
        _selfcheck()
        raise SystemExit(0)
    root = Path(__file__).resolve().parents[1]
    xlsx = sys.argv[1] if len(sys.argv) > 1 else str(Path.home() / "Downloads" / "Список.xlsx")
    jsonl = sys.argv[2] if len(sys.argv) > 2 else str(root / "data" / "nspd-geocode" / "coords.jsonl")
    filled, known = write_coords(xlsx, jsonl)
    print(f"записано {filled} из {known} координат в {xlsx}")
