#!/usr/bin/env python3
"""Импорт RAG.xlsx → data/conclusion_qa.sqlite (FTS5).

ponytail: не LangChain/Chroma из Forpes — в репо уже node:sqlite; FTS5 + топ-k хватает
для «Вопросы по заключению». Эмбеддинги/GigaChat synthesis — опционально поверх retrieval.
"""

from __future__ import annotations

import argparse
import sqlite3
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = ROOT / "data" / "conclusion_qa_rag.xlsx"
DEFAULT_DB = ROOT / "data" / "conclusion_qa.sqlite"

# openpyxl already in requirements-zalog.txt
try:
    import openpyxl
except ImportError:
    print("Нужен openpyxl: npm run zalog:install", file=sys.stderr)
    raise


def _cell(row: tuple, idx: int) -> str:
    if idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).strip()


def _join_answers(row: tuple, start: int = 1) -> str:
    parts = []
    for i in range(start, len(row)):
        t = _cell(row, i)
        if t:
            parts.append(t)
    return "\n---\n".join(parts)


def iter_pairs(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        # ГЛ: ID, Инициатор, Банк, Заемщик, № СЗЗ, ?, Вопрос, Ответ
        if "ГЛ" in wb.sheetnames:
            for i, row in enumerate(wb["ГЛ"].iter_rows(values_only=True)):
                if i == 0:
                    continue
                q, a = _cell(row, 6), _cell(row, 7)
                if q and a:
                    yield "ГЛ", q, a, _cell(row, 0)

        # Для ГЛ: тема + шаблоны ответов
        if "Для ГЛ" in wb.sheetnames:
            for row in wb["Для ГЛ"].iter_rows(values_only=True):
                q, a = _cell(row, 0), _join_answers(row, 1)
                if q and a:
                    yield "Для ГЛ", q, a, ""

        # Осмотры: тема + варианты ответов
        if "Осмотры" in wb.sheetnames:
            for row in wb["Осмотры"].iter_rows(values_only=True):
                q, a = _cell(row, 0), _join_answers(row, 1)
                if q and a:
                    yield "Осмотры", q, a, ""

        # Ошибки экспертов: Суть обращения / Ответ инициатору
        if "Ошибки экспертов" in wb.sheetnames:
            for i, row in enumerate(wb["Ошибки экспертов"].iter_rows(values_only=True)):
                if i == 0:
                    continue
                q, a = _cell(row, 7), _cell(row, 8)
                if q and a:
                    yield "Ошибки экспертов", q, a, _cell(row, 2)

        # Страховки: заметки (вопрос ≈ текст, ответ = тег/вторая колонка или тот же текст)
        if "Страховки" in wb.sheetnames:
            for row in wb["Страховки"].iter_rows(values_only=True):
                q = _cell(row, 0)
                if len(q) < 20:
                    continue
                a = _join_answers(row, 1) or q
                yield "Страховки", q, a, ""
    finally:
        wb.close()


def build(xlsx_path: Path, db_path: Path) -> int:
    if not xlsx_path.is_file():
        raise FileNotFoundError(f"Нет файла: {xlsx_path}")

    db_path.parent.mkdir(parents=True, exist_ok=True)
    if db_path.exists():
        db_path.unlink()

    pairs = list(iter_pairs(xlsx_path))
    con = sqlite3.connect(db_path)
    try:
        con.execute("PRAGMA journal_mode = DELETE;")
        con.executescript(
            """
            CREATE TABLE qa (
                id INTEGER PRIMARY KEY,
                sheet TEXT NOT NULL,
                question TEXT NOT NULL,
                answer TEXT NOT NULL,
                ref_id TEXT
            );
            CREATE VIRTUAL TABLE qa_fts USING fts5(
                question,
                answer,
                content='qa',
                content_rowid='id',
                tokenize='unicode61 remove_diacritics 2'
            );
            """
        )
        insert = con.execute
        for sheet, q, a, ref in pairs:
            cur = insert(
                "INSERT INTO qa(sheet, question, answer, ref_id) VALUES (?, ?, ?, ?)",
                (sheet, q, a, ref or None),
            )
            rowid = cur.lastrowid
            insert(
                "INSERT INTO qa_fts(rowid, question, answer) VALUES (?, ?, ?)",
                (rowid, q, a),
            )
        con.commit()
    finally:
        con.close()
    return len(pairs)


def main() -> int:
    p = argparse.ArgumentParser(description="Сборка FTS5-индекса conclusion Q&A")
    p.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    p.add_argument("--db", type=Path, default=DEFAULT_DB)
    args = p.parse_args()
    n = build(args.xlsx, args.db)
    print(f"OK: {n} пар → {args.db}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
