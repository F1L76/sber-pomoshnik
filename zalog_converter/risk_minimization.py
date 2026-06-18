"""Справочник минимизации рисков из XLSX (столбец C по номеру риска)."""

from __future__ import annotations

import re
from functools import lru_cache
from pathlib import Path

import openpyxl

from .pdf_extract import ConclusionRisk

_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_RAG_XLSX = _ROOT / "data" / "risk_minimization_rag.xlsx"

_RISK_NUM_RE = re.compile(r"^(\d+(?:\.\d+)*|[А-ЯA-ZЁ]+_[\d.]+)")


def _parse_risk_number_from_question(text: str) -> str:
    match = _RISK_NUM_RE.match(text.strip())
    if not match:
        return ""
    return match.group(1).rstrip(".")


@lru_cache(maxsize=4)
def load_minimization_index(xlsx_path: str) -> dict[str, str]:
    path = Path(xlsx_path)
    if not path.is_file():
        return {}

    index: dict[str, str] = {}
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in workbook.sheetnames:
            for row in workbook[sheet_name].iter_rows(values_only=True):
                if len(row) < 3:
                    continue
                question = str(row[1] or "").strip()
                answer = str(row[2] or "").strip()
                if not question or not answer:
                    continue
                number = _parse_risk_number_from_question(question)
                if number:
                    index[number] = answer
    finally:
        workbook.close()
    return index


def lookup_minimization(risk_number: str, xlsx_path: Path | None = None) -> str:
    if not risk_number:
        return "—"
    path = xlsx_path or DEFAULT_RAG_XLSX
    index = load_minimization_index(str(path))
    normalized = risk_number.strip().rstrip(".")
    return index.get(normalized) or index.get(f"{normalized}.") or "—"


def enrich_risks_with_minimization(
    risks: list[ConclusionRisk],
    xlsx_path: Path | None = None,
) -> None:
    path = xlsx_path or DEFAULT_RAG_XLSX
    index = load_minimization_index(str(path))
    for risk in risks:
        if not risk.risk_number:
            # Минимизация уже могла быть извлечена из текста PDF (краткая форма АСЗ)
            continue
        normalized = risk.risk_number.strip().rstrip(".")
        risk.minimization = index.get(normalized) or index.get(f"{normalized}.") or "—"
