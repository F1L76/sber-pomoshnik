"""Чтение листа XLSX с учётом объединённых ячеек."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _fill_merged_values(sheet: Worksheet, rows: list[list[Any]]) -> int:
    """Подставляет значение верхней левой ячейки во все ячейки объединённого диапазона."""
    merged_count = len(sheet.merged_cells.ranges)
    if merged_count == 0:
        return 0

    for merge_range in sheet.merged_cells.ranges:
        top_left = sheet.cell(merge_range.min_row, merge_range.min_col).value
        for row_idx in range(merge_range.min_row, merge_range.max_row + 1):
            list_idx = row_idx - 1
            if list_idx < 0 or list_idx >= len(rows):
                continue
            row = rows[list_idx]
            for col_idx in range(merge_range.min_col, merge_range.max_col + 1):
                cell_idx = col_idx - 1
                if cell_idx < 0:
                    continue
                if cell_idx >= len(row):
                    row.extend([None] * (cell_idx - len(row) + 1))
                current = row[cell_idx]
                if current is None or str(current).strip() == "":
                    row[cell_idx] = top_left
    return merged_count


def read_xlsx_sheet_rows(source: bytes | str) -> tuple[str, list[list[Any]], int]:
    """
    Возвращает (имя листа, строки с заполненными merged cells, число объединённых диапазонов).
    read_only=False — иначе merged_cells недоступны.
    """
    if isinstance(source, str):
        workbook = load_workbook(source, read_only=False, data_only=True)
    else:
        workbook = load_workbook(BytesIO(source), read_only=False, data_only=True)

    try:
        sheet_name = workbook.sheetnames[0]
        sheet = workbook[sheet_name]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        merged_regions_count = _fill_merged_values(sheet, rows)
        return sheet_name, rows, merged_regions_count
    finally:
        workbook.close()
